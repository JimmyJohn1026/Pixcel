from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image
# import multiprocessing

# create excel file
wb = Workbook()
ws = wb.active


# convert number to column
def column_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def paint_row(y, x_max, pix):
    x = 1
    ws.row_dimensions[y].height = 11.25

    while x <= x_max:
        # change column dimensions
        ws.column_dimensions[column_string(x)].width = 2.14

        # get color
        c_hex = "%02x%02x%02x" % pix[x - 1, y - 1]
        c = PatternFill(start_color=c_hex, end_color=c_hex, fill_type="solid")

        # fill color
        cell = "{}{}".format(column_string(x), y)
        ws[cell].fill = c
        x += 1


def main():
    # load the image
    file = input("Image filename (it must be inside this project folder): ")
    im = Image.open(file)
    pix = im.load()

    # get export name
    name = input("Excel filename (.xlsx is added for you): ")

    # image vars
    x_max = im.size[0]  # x pixels
    y_max = im.size[1]  # y pixels

    # paint
    for i in range(y_max):
        paint_row(i + 1, x_max, pix)

    # save the file
    wb.save(name + ".xlsx")


if __name__ == "__main__":
    main()
